"""Image conversion logic for Excel pixel art."""

from __future__ import annotations

from pathlib import Path

from openpyxl import Workbook
from openpyxl.styles import Alignment, Font, PatternFill
from openpyxl.worksheet.worksheet import Worksheet
from PIL import Image, ImageOps

from .canvas import CANVAS_PRESETS, FIT_MODES, ORIENTATIONS, CanvasPreset


def image_to_excel(
    image_path: str | Path,
    output_path: str | Path,
    max_size: int = 32,
    cell_size: float = 3.0,
    color_count: int = 24,
    canvas_size: str | None = None,
    resolution: tuple[int, int] | None = None,
    orientation: str = "auto",
    fit: str = "contain",
    background_color: str = "FFFFFF",
) -> Path:
    """Convert an image into an Excel workbook with one colored cell per pixel."""
    image_path = Path(image_path)
    output_path = Path(output_path)
    canvas_preset = _get_canvas_preset(canvas_size)
    orientation = orientation.lower()
    fit = fit.lower()
    background_color = _normalize_hex_color(background_color)
    if max_size < 1:
        raise ValueError("max_size must be at least 1")
    if cell_size <= 0:
        raise ValueError("cell_size must be greater than 0")
    if color_count < 2 or color_count > 256:
        raise ValueError("color_count must be between 2 and 256")
    if resolution is not None:
        _validate_resolution(resolution)
    if orientation not in ORIENTATIONS:
        raise ValueError(f"orientation must be one of: {', '.join(sorted(ORIENTATIONS))}")
    if fit not in FIT_MODES:
        raise ValueError(f"fit must be one of: {', '.join(sorted(FIT_MODES))}")
    if not image_path.exists():
        raise FileNotFoundError(f"Input image not found: {image_path}")

    source_image = Image.open(image_path).convert("RGBA")
    workbook = Workbook()
    digital_image = _prepare_canvas_image(
        source_image,
        canvas_preset=canvas_preset,
        max_size=max_size,
        resolution=resolution,
        orientation=orientation,
        fit=fit,
        background_color=background_color,
    )
    digital_image = _reduce_colors(digital_image, color_count)
    digital_width, digital_height = digital_image.size
    digital_pixels = digital_image.load()
    digital_palette, digital_color_counts = _build_palette(
        digital_pixels,
        digital_width,
        digital_height,
    )

    reference_sheet = workbook.active
    reference_sheet.title = "Reference"
    _write_reference_sheet(reference_sheet, digital_pixels, digital_width, digital_height, cell_size)
    _configure_page(reference_sheet, canvas_preset, orientation, digital_image.size)

    template_sheet = workbook.create_sheet("Template")
    _write_template_sheet(
        template_sheet,
        digital_pixels,
        digital_palette,
        digital_width,
        digital_height,
        cell_size,
    )
    _configure_page(template_sheet, canvas_preset, orientation, digital_image.size)

    index_sheet = workbook.create_sheet("Color Index")
    _write_color_index_sheet(index_sheet, digital_palette, digital_color_counts)

    output_path.parent.mkdir(parents=True, exist_ok=True)
    workbook.save(output_path)
    return output_path


def _prepare_canvas_image(
    source_image: Image.Image,
    canvas_preset: CanvasPreset | None,
    max_size: int,
    resolution: tuple[int, int] | None,
    orientation: str,
    fit: str,
    background_color: str,
) -> Image.Image:
    image = source_image.copy()
    if resolution is not None:
        return _fit_image_to_canvas(
            image,
            size=resolution,
            fit=fit,
            background_color=background_color,
        )
    if canvas_preset is None:
        image.thumbnail((max_size, max_size), Image.Resampling.LANCZOS)
        return image

    canvas_size = _canvas_dimensions(
        canvas_preset,
        max_size=max_size,
        orientation=orientation,
        image_size=image.size,
    )
    return _fit_image_to_canvas(
        image,
        size=canvas_size,
        fit=fit,
        background_color=background_color,
    )


def _build_palette(pixels, width: int, height: int) -> tuple[dict[str, int], dict[str, int]]:
    counts: dict[str, int] = {}
    for row in range(height):
        for column in range(width):
            red, green, blue, alpha = pixels[column, row]
            if alpha == 0:
                continue

            color = f"{red:02X}{green:02X}{blue:02X}"
            counts[color] = counts.get(color, 0) + 1

    ordered_colors = sorted(counts, key=lambda color: (-counts[color], color))
    palette = {color: index for index, color in enumerate(ordered_colors, start=1)}
    return palette, counts


def _write_reference_sheet(sheet: Worksheet, pixels, width: int, height: int, cell_size: float) -> None:
    sheet.sheet_view.showGridLines = False
    fills: dict[str, PatternFill] = {}
    _set_pixel_dimensions(sheet, width, height, cell_size)

    for row in range(height):
        for column in range(width):
            red, green, blue, alpha = pixels[column, row]
            if alpha == 0:
                continue

            color = f"{red:02X}{green:02X}{blue:02X}"
            fill = fills.setdefault(color, PatternFill(fill_type="solid", fgColor=color))
            sheet.cell(row=row + 1, column=column + 1).fill = fill


def _write_template_sheet(
    sheet: Worksheet,
    pixels,
    palette: dict[str, int],
    width: int,
    height: int,
    cell_size: float,
) -> None:
    sheet.sheet_view.showGridLines = True
    sheet.print_options.gridLines = False
    _set_pixel_dimensions(sheet, width, height, cell_size)

    number_font = Font(color="A6A6A6", size=6)
    center = Alignment(horizontal="center", vertical="center", shrink_to_fit=True)

    for row in range(height):
        for column in range(width):
            cell = sheet.cell(row=row + 1, column=column + 1)
            cell.font = number_font
            cell.alignment = center

            red, green, blue, alpha = pixels[column, row]
            if alpha == 0:
                continue

            color = f"{red:02X}{green:02X}{blue:02X}"
            cell.value = palette[color]


def _write_color_index_sheet(sheet: Worksheet, palette: dict[str, int], color_counts: dict[str, int]) -> None:
    sheet.freeze_panes = "A2"
    sheet.column_dimensions["A"].width = 10
    sheet.column_dimensions["B"].width = 14
    sheet.column_dimensions["C"].width = 12
    sheet.column_dimensions["D"].width = 14

    headers = ["Number", "Hex Code", "Swatch", "Cells"]
    for column, header in enumerate(headers, start=1):
        cell = sheet.cell(row=1, column=column)
        cell.value = header
        cell.font = Font(bold=True)
        cell.alignment = Alignment(horizontal="center")

    for color, number in sorted(palette.items(), key=lambda item: item[1]):
        row = number + 1
        sheet.cell(row=row, column=1).value = number
        sheet.cell(row=row, column=2).value = f"#{color}"
        sheet.cell(row=row, column=3).fill = PatternFill(fill_type="solid", fgColor=color)
        sheet.cell(row=row, column=4).value = color_counts[color]


def _set_pixel_dimensions(sheet: Worksheet, width: int, height: int, cell_size: float) -> None:
    for row in range(height):
        sheet.row_dimensions[row + 1].height = cell_size * 6
    for column in range(width):
        sheet.column_dimensions[_column_name(column + 1)].width = cell_size


def _get_canvas_preset(canvas_size: str | None) -> CanvasPreset | None:
    if canvas_size is None or canvas_size == "":
        return None

    key = canvas_size.lower()
    if key not in CANVAS_PRESETS:
        choices = ", ".join(sorted(CANVAS_PRESETS))
        raise ValueError(f"canvas_size must be one of: {choices}")
    return CANVAS_PRESETS[key]


def _validate_resolution(resolution: tuple[int, int]) -> None:
    width, height = resolution
    if width < 1 or height < 1:
        raise ValueError("resolution width and height must be at least 1")
    if width > 2000 or height > 2000:
        raise ValueError("resolution width and height must be at most 2000")


def _canvas_dimensions(
    preset: CanvasPreset,
    max_size: int,
    orientation: str,
    image_size: tuple[int, int],
) -> tuple[int, int]:
    width_mm, height_mm = _oriented_size(preset, orientation, image_size)
    if width_mm >= height_mm:
        width = max_size
        height = max(1, round(max_size * height_mm / width_mm))
    else:
        height = max_size
        width = max(1, round(max_size * width_mm / height_mm))
    return width, height


def _oriented_size(
    preset: CanvasPreset,
    orientation: str,
    image_size: tuple[int, int],
) -> tuple[float, float]:
    width_mm = preset.width_mm
    height_mm = preset.height_mm
    if orientation == "auto":
        orientation = "landscape" if image_size[0] >= image_size[1] else "portrait"

    if orientation == "landscape" and width_mm < height_mm:
        return height_mm, width_mm
    if orientation == "portrait" and width_mm > height_mm:
        return height_mm, width_mm
    return width_mm, height_mm


def _fit_image_to_canvas(
    image: Image.Image,
    size: tuple[int, int],
    fit: str,
    background_color: str,
) -> Image.Image:
    if fit == "cover":
        return ImageOps.fit(image, size, method=Image.Resampling.LANCZOS, centering=(0.5, 0.5))

    fitted = image.copy()
    fitted.thumbnail(size, Image.Resampling.LANCZOS)
    canvas = Image.new("RGBA", size, f"#{background_color}")
    left = (size[0] - fitted.width) // 2
    top = (size[1] - fitted.height) // 2
    canvas.alpha_composite(fitted, dest=(left, top))
    return canvas


def _reduce_colors(image: Image.Image, color_count: int) -> Image.Image:
    if image.mode != "RGBA":
        image = image.convert("RGBA")

    alpha = image.getchannel("A")
    opaque = Image.new("RGBA", image.size, "#FFFFFF")
    opaque.alpha_composite(image)
    quantized = opaque.convert("RGB").quantize(colors=color_count, method=Image.Quantize.MEDIANCUT)
    reduced = quantized.convert("RGBA")
    reduced.putalpha(alpha)
    return reduced


def _configure_page(
    sheet: Worksheet,
    preset: CanvasPreset | None,
    orientation: str,
    image_size: tuple[int, int],
) -> None:
    sheet.page_setup.fitToWidth = 1
    sheet.page_setup.fitToHeight = 1
    sheet.sheet_properties.pageSetUpPr.fitToPage = True
    sheet.page_margins.left = 0.25
    sheet.page_margins.right = 0.25
    sheet.page_margins.top = 0.25
    sheet.page_margins.bottom = 0.25

    if preset is None:
        sheet.page_setup.orientation = "landscape" if image_size[0] >= image_size[1] else "portrait"
        return

    if preset.excel_paper_size is not None:
        sheet.page_setup.paperSize = preset.excel_paper_size
    if orientation == "auto":
        orientation = "landscape" if image_size[0] >= image_size[1] else "portrait"
    sheet.page_setup.orientation = orientation


def _normalize_hex_color(color: str) -> str:
    normalized = color.strip().lstrip("#").upper()
    if not re_match_hex_color(normalized):
        raise ValueError("background_color must be a 6-digit hex color")
    return normalized


def re_match_hex_color(color: str) -> bool:
    if len(color) != 6:
        return False
    return all(character in "0123456789ABCDEF" for character in color)


def _column_name(index: int) -> str:
    name = ""
    while index:
        index, remainder = divmod(index - 1, 26)
        name = chr(65 + remainder) + name
    return name
