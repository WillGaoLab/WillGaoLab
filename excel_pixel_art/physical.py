"""Independent Physical Layer workbook generation."""

from __future__ import annotations

import argparse
import zipfile
from pathlib import Path

from openpyxl import Workbook
from openpyxl.styles import Alignment, Font, PatternFill
from openpyxl.worksheet.pagebreak import Break
from openpyxl.worksheet.worksheet import Worksheet
from PIL import Image, ImageOps

from .canvas import CANVAS_PRESETS, FIT_MODES, ORIENTATIONS, CanvasPreset
from .poster_print import write_vector_poster_pdf

MATERIAL_PALETTE_REFERENCES = {
    "adaptive": None,
    "lego": "https://www.bricklink.com/catalogColors.asp",
    "liquitex_basics_24": "https://www.michaels.com/product/liquitex-basics-acrylic-24-color-paint-set-10268659",
}

# Screen approximations for matching images to purchasable physical materials.
LEGO_COLORS = [
    ("White", "FFFFFF"), ("Light Bluish Gray", "A0A5A9"), ("Dark Bluish Gray", "6C6E68"),
    ("Black", "05131D"), ("Dark Red", "720E0F"), ("Red", "C91A09"),
    ("Coral", "FF698F"), ("Dark Brown", "352100"), ("Reddish Brown", "582A12"),
    ("Dark Tan", "958A73"), ("Tan", "E4CD9E"), ("Light Nougat", "F6D7B3"),
    ("Medium Nougat", "AA7D55"), ("Dark Orange", "A95500"), ("Orange", "FE8A18"),
    ("Bright Light Orange", "F8BB3D"), ("Yellow", "F2CD37"), ("Bright Light Yellow", "FFF03A"),
    ("Lime", "BBE90B"), ("Olive Green", "9B9A5A"), ("Dark Green", "184632"),
    ("Green", "237841"), ("Bright Green", "4B9F4A"), ("Sand Green", "A0BCAC"),
    ("Dark Turquoise", "008F9B"), ("Aqua", "B3D7D1"), ("Dark Blue", "0A3463"),
    ("Blue", "0055BF"), ("Dark Azure", "078BC9"), ("Medium Azure", "36AEBF"),
    ("Medium Blue", "5A93DB"), ("Bright Light Blue", "9FC3E9"), ("Sand Blue", "6074A1"),
    ("Dark Purple", "3F3691"), ("Purple", "81007B"), ("Medium Lavender", "AC78BA"),
    ("Lavender", "E1D5ED"), ("Magenta", "923978"), ("Dark Pink", "C870A0"),
    ("Bright Pink", "E4ADC8"),
]

LIQUITEX_BASICS_24 = [
    ("Bright Aqua Green", "00A99D"), ("Burnt Sienna", "8A3F2B"),
    ("Burnt Umber", "4B352A"), ("Cadmium Orange Hue", "F26A21"),
    ("Cadmium Red Deep Hue", "9E1B32"), ("Cadmium Red Medium Hue", "D9272E"),
    ("Cadmium Yellow Deep Hue", "F6A800"), ("Cadmium Yellow Medium Hue", "F5D000"),
    ("Dioxazine Purple", "3E2465"), ("Iridescent Gold", "B08D35"),
    ("Iridescent Silver", "B7B7B7"), ("Ivory Black", "1C1B1A"),
    ("Mars Black", "101010"), ("Medium Magenta", "A62974"),
    ("Naphthol Crimson", "B51E3A"), ("Permanent Green Light", "5AAE3A"),
    ("Permanent Hooker's Green", "24543D"), ("Permanent Light Blue", "62A9D8"),
    ("Phthalo Blue", "0B4F8A"), ("Primary Blue", "1F4E9D"),
    ("Primary Yellow", "F4D400"), ("Titanium White", "F7F6F0"),
    ("Ultramarine Blue", "273D8F"), ("Unbleached Titanium", "D2B48C"),
]

MATERIAL_PALETTES = {
    "lego": LEGO_COLORS,
    "liquitex_basics_24": LIQUITEX_BASICS_24,
}


def image_to_physical_excel(
    image_path: str | Path,
    output_path: str | Path,
    max_size: int = 32,
    cell_size: float = 3.0,
    material_color_count: int = 24,
    canvas_size: str | None = "a4",
    resolution: tuple[int, int] | None = (32, 32),
    orientation: str = "auto",
    fit: str = "contain",
    background_color: str = "FFFFFF",
    poster_pages: tuple[int, int] = (2, 2),
    generate_color_masks: bool = False,
    max_color_masks: int | None = None,
    palette_mode: str = "adaptive",
) -> Path:
    """Generate a standalone Physical Layer print workbook."""
    image_path = Path(image_path)
    output_path = Path(output_path)
    canvas_preset = _get_canvas_preset(canvas_size)
    orientation = orientation.lower()
    fit = fit.lower()
    background_color = _normalize_hex_color(background_color)

    if max_size < 1:
        raise ValueError("max_size must be at least 1")
    if cell_size <= 0:
        raise ValueError("cell_size must be greater than 0")
    if palette_mode not in {"adaptive", *MATERIAL_PALETTES}:
        raise ValueError("palette_mode must be adaptive, lego, or liquitex_basics_24")
    if material_color_count < 2 or material_color_count > 256:
        raise ValueError("material_color_count must be between 2 and 256")
    if resolution is not None:
        _validate_resolution(resolution)
    if orientation not in ORIENTATIONS:
        raise ValueError(f"orientation must be one of: {', '.join(sorted(ORIENTATIONS))}")
    if fit not in FIT_MODES:
        raise ValueError(f"fit must be one of: {', '.join(sorted(FIT_MODES))}")
    _validate_print_options(poster_pages, max_color_masks)
    if not image_path.exists():
        raise FileNotFoundError(f"Input image not found: {image_path}")

    physical_image, material_names = _prepare_physical_image(
        image_path, canvas_preset, max_size, resolution, orientation, fit,
        background_color, material_color_count, palette_mode,
    )
    width, height = physical_image.size
    pixels = physical_image.load()
    palette, color_counts = _build_palette(pixels, width, height)

    workbook = Workbook()
    reference_sheet = workbook.active
    reference_sheet.title = "Print Reference"
    _write_reference_sheet(reference_sheet, pixels, width, height, cell_size)
    _configure_page(reference_sheet, canvas_preset, orientation, physical_image.size)
    _configure_print_mode(reference_sheet, width, height, poster_pages)

    template_sheet = workbook.create_sheet("Print Template")
    _write_template_sheet(template_sheet, pixels, palette, width, height, cell_size)
    _configure_page(template_sheet, canvas_preset, orientation, physical_image.size)
    _configure_print_mode(template_sheet, width, height, poster_pages)

    material_sheet = workbook.create_sheet("Material Palette")
    _write_color_index_sheet(
        material_sheet,
        palette,
        color_counts,
        material_names=material_names,
        reference=MATERIAL_PALETTE_REFERENCES[palette_mode],
    )

    if generate_color_masks:
        _write_color_mask_sheets(
            workbook=workbook,
            pixels=pixels,
            palette=palette,
            width=width,
            height=height,
            cell_size=cell_size,
            canvas_preset=canvas_preset,
            orientation=orientation,
            image_size=physical_image.size,
            poster_pages=poster_pages,
            max_color_masks=max_color_masks,
        )

    output_path.parent.mkdir(parents=True, exist_ok=True)
    workbook.save(output_path)
    return output_path


def image_to_physical_pdf(
    image_path: str | Path,
    output_path: str | Path,
    **options,
) -> Path:
    """Generate a colored, numbered, and coordinated tiled printable PDF."""
    image_path = Path(image_path)
    output_path = Path(output_path)
    poster_pages = options.pop("poster_pages", (2, 2))
    image, _ = _prepare_physical_from_options(image_path, options)
    return write_vector_poster_pdf(image, output_path, poster_pages)


def image_to_physical_masks(
    image_path: str | Path,
    output_path: str | Path,
    max_color_masks: int | None = None,
    **options,
) -> Path:
    """Generate a ZIP containing numbered black-and-white material masks."""
    image_path = Path(image_path)
    output_path = Path(output_path)
    image, names = _prepare_physical_from_options(image_path, options)
    pixels = image.load()
    width, height = image.size
    palette, _ = _build_palette(pixels, width, height)
    colors = sorted(palette, key=palette.get)
    if max_color_masks is not None:
        colors = colors[:max_color_masks]

    output_path.parent.mkdir(parents=True, exist_ok=True)
    with zipfile.ZipFile(output_path, "w", compression=zipfile.ZIP_DEFLATED) as archive:
        for color in colors:
            number = palette[color]
            mask = Image.new("1", image.size, 1)
            mask_pixels = mask.load()
            for row in range(height):
                for column in range(width):
                    red, green, blue, alpha = pixels[column, row]
                    if alpha > 0 and f"{red:02X}{green:02X}{blue:02X}" == color:
                        mask_pixels[column, row] = 0
            name = names.get(color, color).replace("/", "-")
            temporary = output_path.parent / f"{number:03d}_{name}.png"
            mask.resize((width * 8, height * 8), Image.Resampling.NEAREST).save(temporary)
            archive.write(temporary, temporary.name)
            temporary.unlink()
    return output_path


def _prepare_physical_from_options(image_path: Path, options: dict) -> tuple[Image.Image, dict[str, str]]:
    canvas_preset = _get_canvas_preset(options.get("canvas_size", "a4"))
    return _prepare_physical_image(
        image_path,
        canvas_preset,
        options.get("max_size", 32),
        options.get("resolution", (32, 32)),
        options.get("orientation", "auto"),
        options.get("fit", "contain"),
        _normalize_hex_color(options.get("background_color", "FFFFFF")),
        options.get("material_color_count", 24),
        options.get("palette_mode", "adaptive"),
    )


def _prepare_physical_image(
    image_path: Path,
    canvas_preset: CanvasPreset | None,
    max_size: int,
    resolution: tuple[int, int] | None,
    orientation: str,
    fit: str,
    background_color: str,
    material_color_count: int,
    palette_mode: str,
) -> tuple[Image.Image, dict[str, str]]:
    source_image = Image.open(image_path).convert("RGBA")
    image = _prepare_canvas_image(
        source_image, canvas_preset, max_size, resolution, orientation, fit, background_color,
    )
    if palette_mode == "adaptive":
        return _reduce_colors(image, material_color_count), {}
    return _match_fixed_palette(image, MATERIAL_PALETTES[palette_mode])


def _match_fixed_palette(image: Image.Image, materials: list[tuple[str, str]]) -> tuple[Image.Image, dict[str, str]]:
    output = Image.new("RGBA", image.size)
    source = image.load()
    target = output.load()
    colors = [
        (name, value, tuple(bytes.fromhex(value)), _rgb_to_lab(tuple(bytes.fromhex(value))))
        for name, value in materials
    ]
    for row in range(image.height):
        for column in range(image.width):
            red, green, blue, alpha = source[column, row]
            source_lab = _rgb_to_lab((red, green, blue))
            name, value, rgb, _ = min(
                colors,
                key=lambda item: sum((source_lab[index] - item[3][index]) ** 2 for index in range(3)),
            )
            target[column, row] = (*rgb, alpha)
    return output, {value: name for name, value in materials}


def _rgb_to_lab(rgb: tuple[int, int, int]) -> tuple[float, float, float]:
    linear = []
    for value in rgb:
        channel = value / 255
        linear.append(channel / 12.92 if channel <= 0.04045 else ((channel + 0.055) / 1.055) ** 2.4)
    red, green, blue = linear
    x = (red * 0.4124 + green * 0.3576 + blue * 0.1805) / 0.95047
    y = red * 0.2126 + green * 0.7152 + blue * 0.0722
    z = (red * 0.0193 + green * 0.1192 + blue * 0.9505) / 1.08883

    def transform(value: float) -> float:
        return value ** (1 / 3) if value > 0.008856 else 7.787 * value + 16 / 116

    x, y, z = transform(x), transform(y), transform(z)
    return 116 * y - 16, 500 * (x - y), 200 * (y - z)


def _split_image_pages(image: Image.Image, poster_pages: tuple[int, int]) -> list[Image.Image]:
    pages_wide, pages_tall = poster_pages
    pages = []
    for page_y in range(pages_tall):
        for page_x in range(pages_wide):
            left = round(image.width * page_x / pages_wide)
            right = round(image.width * (page_x + 1) / pages_wide)
            top = round(image.height * page_y / pages_tall)
            bottom = round(image.height * (page_y + 1) / pages_tall)
            pages.append(image.crop((left, top, right, bottom)).resize((1800, 2400), Image.Resampling.NEAREST))
    return pages


def _write_color_mask_sheets(
    workbook: Workbook,
    pixels,
    palette: dict[str, int],
    width: int,
    height: int,
    cell_size: float,
    canvas_preset: CanvasPreset | None,
    orientation: str,
    image_size: tuple[int, int],
    poster_pages: tuple[int, int],
    max_color_masks: int | None,
) -> None:
    colors = sorted(palette, key=palette.get)
    if max_color_masks is not None:
        colors = colors[:max_color_masks]

    mask_fill = PatternFill(fill_type="solid", fgColor="000000")
    for color in colors:
        number = palette[color]
        sheet = workbook.create_sheet(f"Material Mask {number:03d}")
        sheet.sheet_view.showGridLines = False
        sheet.print_options.gridLines = False
        sheet.sheet_properties.tabColor = color
        sheet.oddHeader.center.text = f"Color {number} - #{color}"
        _set_pixel_dimensions(sheet, width, height, cell_size)

        for row in range(height):
            for column in range(width):
                red, green, blue, alpha = pixels[column, row]
                pixel_color = f"{red:02X}{green:02X}{blue:02X}"
                if alpha > 0 and pixel_color == color:
                    sheet.cell(row=row + 1, column=column + 1).fill = mask_fill

        _configure_page(sheet, canvas_preset, orientation, image_size)
        _configure_print_mode(sheet, width, height, poster_pages)


def _prepare_canvas_image(
    source_image: Image.Image,
    canvas_preset: CanvasPreset | None,
    max_size: int,
    resolution: tuple[int, int] | None,
    orientation: str,
    fit: str,
    background_color: str,
) -> Image.Image:
    image = source_image.copy()
    if resolution is not None:
        return _fit_image_to_canvas(image, resolution, fit, background_color)
    if canvas_preset is None:
        image.thumbnail((max_size, max_size), Image.Resampling.LANCZOS)
        return image

    canvas_size = _canvas_dimensions(canvas_preset, max_size, orientation, image.size)
    return _fit_image_to_canvas(image, canvas_size, fit, background_color)


def _build_palette(pixels, width: int, height: int) -> tuple[dict[str, int], dict[str, int]]:
    counts: dict[str, int] = {}
    for row in range(height):
        for column in range(width):
            red, green, blue, alpha = pixels[column, row]
            if alpha == 0:
                continue
            color = f"{red:02X}{green:02X}{blue:02X}"
            counts[color] = counts.get(color, 0) + 1

    ordered_colors = sorted(counts, key=lambda color: (-counts[color], color))
    return {color: index for index, color in enumerate(ordered_colors, start=1)}, counts


def _write_reference_sheet(sheet: Worksheet, pixels, width: int, height: int, cell_size: float) -> None:
    sheet.sheet_view.showGridLines = False
    fills: dict[str, PatternFill] = {}
    _set_pixel_dimensions(sheet, width, height, cell_size)
    for row in range(height):
        for column in range(width):
            red, green, blue, alpha = pixels[column, row]
            if alpha == 0:
                continue
            color = f"{red:02X}{green:02X}{blue:02X}"
            sheet.cell(row=row + 1, column=column + 1).fill = fills.setdefault(
                color,
                PatternFill(fill_type="solid", fgColor=color),
            )


def _write_template_sheet(
    sheet: Worksheet,
    pixels,
    palette: dict[str, int],
    width: int,
    height: int,
    cell_size: float,
) -> None:
    sheet.sheet_view.showGridLines = True
    sheet.print_options.gridLines = False
    _set_pixel_dimensions(sheet, width, height, cell_size)
    number_font = Font(color="A6A6A6", size=6)
    center = Alignment(horizontal="center", vertical="center", shrink_to_fit=True)

    for row in range(height):
        for column in range(width):
            cell = sheet.cell(row=row + 1, column=column + 1)
            cell.font = number_font
            cell.alignment = center
            red, green, blue, alpha = pixels[column, row]
            if alpha > 0:
                cell.value = palette[f"{red:02X}{green:02X}{blue:02X}"]


def _write_color_index_sheet(
    sheet: Worksheet,
    palette: dict[str, int],
    color_counts: dict[str, int],
    material_names: dict[str, str] | None = None,
    reference: str | None = None,
) -> None:
    material_names = material_names or {}
    sheet.freeze_panes = "A2"
    sheet.column_dimensions["A"].width = 10
    sheet.column_dimensions["B"].width = 14
    sheet.column_dimensions["C"].width = 12
    sheet.column_dimensions["D"].width = 14
    sheet.column_dimensions["E"].width = 28

    for column, header in enumerate(["Number", "Hex Code", "Swatch", "Cells", "Material"], start=1):
        cell = sheet.cell(row=1, column=column)
        cell.value = header
        cell.font = Font(bold=True)
        cell.alignment = Alignment(horizontal="center")

    for color, number in sorted(palette.items(), key=lambda item: item[1]):
        row = number + 1
        sheet.cell(row=row, column=1).value = number
        sheet.cell(row=row, column=2).value = f"#{color}"
        sheet.cell(row=row, column=3).fill = PatternFill(fill_type="solid", fgColor=color)
        sheet.cell(row=row, column=4).value = color_counts[color]
        sheet.cell(row=row, column=5).value = material_names.get(color, "Adaptive Color")
    if reference:
        sheet["G1"] = "Reference"
        sheet["G2"] = reference


def _set_pixel_dimensions(sheet: Worksheet, width: int, height: int, cell_size: float) -> None:
    for row in range(height):
        sheet.row_dimensions[row + 1].height = cell_size * 6
    for column in range(width):
        sheet.column_dimensions[_column_name(column + 1)].width = cell_size


def _get_canvas_preset(canvas_size: str | None) -> CanvasPreset | None:
    if canvas_size is None or canvas_size == "":
        return None
    key = canvas_size.lower()
    if key not in CANVAS_PRESETS:
        raise ValueError(f"canvas_size must be one of: {', '.join(sorted(CANVAS_PRESETS))}")
    return CANVAS_PRESETS[key]


def _validate_resolution(resolution: tuple[int, int]) -> None:
    width, height = resolution
    if width < 1 or height < 1:
        raise ValueError("resolution width and height must be at least 1")
    if width > 2000 or height > 2000:
        raise ValueError("resolution width and height must be at most 2000")


def _validate_print_options(poster_pages: tuple[int, int], max_color_masks: int | None) -> None:
    pages_wide, pages_tall = poster_pages
    if pages_wide < 1 or pages_tall < 1:
        raise ValueError("poster page dimensions must be at least 1")
    if pages_wide > 20 or pages_tall > 20:
        raise ValueError("poster page dimensions must be at most 20")
    if max_color_masks is not None and (max_color_masks < 1 or max_color_masks > 256):
        raise ValueError("max_color_masks must be between 1 and 256")


def _canvas_dimensions(
    preset: CanvasPreset,
    max_size: int,
    orientation: str,
    image_size: tuple[int, int],
) -> tuple[int, int]:
    width_mm, height_mm = _oriented_size(preset, orientation, image_size)
    if width_mm >= height_mm:
        return max_size, max(1, round(max_size * height_mm / width_mm))
    return max(1, round(max_size * width_mm / height_mm)), max_size


def _oriented_size(
    preset: CanvasPreset,
    orientation: str,
    image_size: tuple[int, int],
) -> tuple[float, float]:
    width_mm, height_mm = preset.width_mm, preset.height_mm
    if orientation == "auto":
        orientation = "landscape" if image_size[0] >= image_size[1] else "portrait"
    if orientation == "landscape" and width_mm < height_mm:
        return height_mm, width_mm
    if orientation == "portrait" and width_mm > height_mm:
        return height_mm, width_mm
    return width_mm, height_mm


def _fit_image_to_canvas(
    image: Image.Image,
    size: tuple[int, int],
    fit: str,
    background_color: str,
) -> Image.Image:
    if fit == "cover":
        return ImageOps.fit(image, size, method=Image.Resampling.LANCZOS, centering=(0.5, 0.5))
    fitted = image.copy()
    fitted.thumbnail(size, Image.Resampling.LANCZOS)
    canvas = Image.new("RGBA", size, f"#{background_color}")
    canvas.alpha_composite(fitted, dest=((size[0] - fitted.width) // 2, (size[1] - fitted.height) // 2))
    return canvas


def _reduce_colors(image: Image.Image, color_count: int) -> Image.Image:
    if image.mode != "RGBA":
        image = image.convert("RGBA")
    alpha = image.getchannel("A")
    opaque = Image.new("RGBA", image.size, "#FFFFFF")
    opaque.alpha_composite(image)
    reduced = opaque.convert("RGB").quantize(colors=color_count, method=Image.Quantize.MEDIANCUT).convert("RGBA")
    reduced.putalpha(alpha)
    return reduced


def _configure_page(
    sheet: Worksheet,
    preset: CanvasPreset | None,
    orientation: str,
    image_size: tuple[int, int],
) -> None:
    sheet.page_setup.fitToWidth = 1
    sheet.page_setup.fitToHeight = 1
    sheet.sheet_properties.pageSetUpPr.fitToPage = True
    sheet.page_margins.left = 0.25
    sheet.page_margins.right = 0.25
    sheet.page_margins.top = 0.25
    sheet.page_margins.bottom = 0.25
    if preset is None:
        sheet.page_setup.orientation = "landscape" if image_size[0] >= image_size[1] else "portrait"
        return
    if preset.excel_paper_size is not None:
        sheet.page_setup.paperSize = preset.excel_paper_size
    if orientation == "auto":
        orientation = "landscape" if image_size[0] >= image_size[1] else "portrait"
    sheet.page_setup.orientation = orientation


def _configure_print_mode(
    sheet: Worksheet,
    width: int,
    height: int,
    poster_pages: tuple[int, int],
) -> None:
    pages_wide, pages_tall = poster_pages
    sheet.print_area = f"A1:{_column_name(width)}{height}"
    sheet.page_setup.fitToWidth = pages_wide
    sheet.page_setup.fitToHeight = pages_tall
    sheet.sheet_properties.pageSetUpPr.fitToPage = True

    for page in range(1, pages_wide):
        sheet.col_breaks.append(Break(id=round(width * page / pages_wide)))
    for page in range(1, pages_tall):
        sheet.row_breaks.append(Break(id=round(height * page / pages_tall)))


def _normalize_hex_color(color: str) -> str:
    normalized = color.strip().lstrip("#").upper()
    if len(normalized) != 6 or any(character not in "0123456789ABCDEF" for character in normalized):
        raise ValueError("background_color must be a 6-digit hex color")
    return normalized


def _column_name(index: int) -> str:
    name = ""
    while index:
        index, remainder = divmod(index - 1, 26)
        name = chr(65 + remainder) + name
    return name


def parse_resolution(value: str) -> tuple[int, int]:
    normalized = value.lower().replace(" ", "")
    if "x" not in normalized:
        raise argparse.ArgumentTypeError("resolution must use WIDTHxHEIGHT")
    width_text, height_text = normalized.split("x", 1)
    try:
        resolution = int(width_text), int(height_text)
        _validate_resolution(resolution)
    except ValueError as error:
        raise argparse.ArgumentTypeError(str(error)) from error
    return resolution


def build_parser() -> argparse.ArgumentParser:
    parser = argparse.ArgumentParser(description="Generate an independent Physical Layer print workbook.")
    parser.add_argument("image", type=Path)
    parser.add_argument("-o", "--output", type=Path)
    parser.add_argument("--max-size", type=int, default=32)
    parser.add_argument("--cell-size", type=float, default=3.0)
    parser.add_argument("--material-colors", type=int, default=24)
    parser.add_argument("--palette", choices=["adaptive", "lego", "liquitex_basics_24"], default="adaptive")
    parser.add_argument("--canvas-size", choices=sorted(CANVAS_PRESETS), default="a4")
    parser.add_argument("--resolution", type=parse_resolution, default=(32, 32))
    parser.add_argument("--orientation", choices=sorted(ORIENTATIONS), default="auto")
    parser.add_argument("--fit", choices=sorted(FIT_MODES), default="contain")
    parser.add_argument("--background-color", default="FFFFFF")
    parser.add_argument("--pages-wide", type=int, default=2)
    parser.add_argument("--pages-tall", type=int, default=2)
    parser.add_argument("--color-masks", action="store_true")
    parser.add_argument("--max-color-masks", type=int)
    return parser


def main(argv: list[str] | None = None) -> int:
    args = build_parser().parse_args(argv)
    output = args.output or args.image.with_name(f"{args.image.stem}_physical.xlsx")
    try:
        image_to_physical_excel(
            image_path=args.image,
            output_path=output,
            max_size=args.max_size,
            cell_size=args.cell_size,
            material_color_count=args.material_colors,
            palette_mode=args.palette,
            canvas_size=args.canvas_size,
            resolution=args.resolution,
            orientation=args.orientation,
            fit=args.fit,
            background_color=args.background_color,
            poster_pages=(args.pages_wide, args.pages_tall),
            generate_color_masks=args.color_masks,
            max_color_masks=args.max_color_masks,
        )
    except (FileNotFoundError, ValueError) as error:
        print(f"Error: {error}")
        return 1
    print(f"Wrote {output}")
    return 0


if __name__ == "__main__":
    raise SystemExit(main())
